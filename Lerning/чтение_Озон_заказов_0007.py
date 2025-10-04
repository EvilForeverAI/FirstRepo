import openpyxl
input_file_1  = r"F:\Озон TEMP\Наши товары\2025.05.30\Аналитика-графики-по всем уровнем категорий — копия.xlsx"
input_file_2 = r"F:\Озон TEMP\Наши товары\Таблица всех артикулов 465.xlsx"

# Загружаем обе таблицы
wb1 = openpyxl.load_workbook(input_file_1, read_only=True)  # Основная таблица
wb2 = openpyxl.load_workbook(input_file_2, read_only=True)  # Таблица для проверки

sheet1 = wb1.worksheets[0]  # Активный лист в table1
sheet2 = wb2.worksheets[0]  # Активный лист в table2

def find_last_column(sheet):
    last_column = 1  # Минимальный столбец (A)
    for row in sheet.iter_rows():
        for cell in reversed(row):
            if cell.value is not None:
                if cell.column > last_column:  # Нашли столбец правее
                    last_column = cell.column
                break  # Переходим к следующей строке
    return last_column



# Собираем все артикулы из table2 в множество (для быстрого поиска)
articles_table2 = set()
for row in sheet2.iter_rows(min_row=2,values_only=True):
    article = int(row[0])  # Артикулы в 1-м столбце (A) т.е. "0" счет столбцов начинается с 0, строк с 1
    if article is not None:  # хз, за чем, но пусть будет, если ячейка не пустая, то добавляем в общий список
        articles_table2.add(article)
print(articles_table2) # проверка состава списка, все int

# Добавляем новый столбец "Номер Склада" в table1
print(find_last_column(sheet1))
# sheet1.cell(row=1, column=2, value="Наличие на складе")  # Заголовок
#
# # Проверяем каждый артикул в table1
# for idx, row in enumerate(sheet1.iter_rows(min_row=2, values_only=True), start=2):
#     article = str(row[0]).strip()  # Артикул в столбце A (первый столбец)
#     if article in articles_table2:
#         sheet1.cell(row=idx, column=2, value="Есть")  # Отмечаем "Есть"
#     else:
#         sheet1.cell(row=idx, column=2, value="Нет")   # Или оставляем пустым
#
# # Сохраняем изменения
# wb1.save("table1_updated.xlsx")
# print("Готово! Проверьте файл table1_updated.xlsx")
