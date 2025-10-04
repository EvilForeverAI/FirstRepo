import pandas as pd
from datetime import datetime
from openpyxl import load_workbook


def add_new_date_to_excel(existing_file, new_data_file):
    # Текущая дата
    current_date = datetime.now().strftime("%d.%m.%Y")

    # Читаем новые данные
    new_df = pd.read_excel(new_data_file, sheet_name=0, header=2, skiprows=[3])

    # Проверяем нужные столбцы
    if not all(col in new_df.columns for col in ['Кластер', 'Артикул', 'Доступно к продаже']):
        raise ValueError("В новых данных отсутствуют необходимые столбцы")

    # Загружаем существующий файл
    book = load_workbook(existing_file)

    # Обрабатываем каждый лист
    for sheet_name in book.sheetnames:
        # Читаем старые данные
        old_df = pd.read_excel(existing_file, sheet_name=sheet_name)

        # Проверяем, не добавляли ли уже эту дату
        if current_date in old_df.columns:
            continue

        # Фильтруем новые данные по кластеру
        cluster_data = new_df[new_df['Кластер'] == sheet_name]

        # Группируем и суммируем
        new_grouped = cluster_data.groupby('Артикул')['Доступно к продаже'].sum().reset_index()

        # Объединяем со старыми данными
        merged = pd.merge(
            old_df,
            new_grouped,
            on='Артикул',
            how='left'
        ).fillna(0)

        # Переименовываем новый столбец
        merged = merged.rename(columns={'Доступно к продаже': current_date})

        # Сохраняем порядок столбцов
        cols = old_df.columns.tolist() + [current_date]
        merged = merged[cols]

        # Записываем обратно в лист
        with pd.ExcelWriter(existing_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            merged.to_excel(writer, sheet_name=sheet_name, index=False)

    print(f"Файл обновлен. Добавлена дата: {current_date}")

# Пример использования
input_filename = r"F:\Озон TEMP\Эксперименты\stock-report (7).xlsx"  # Замените на ваш файл
existing_file = r"F:\Озон TEMP\Эксперименты\output.xlsx"  # Имя выходного файла

# Использование
add_new_date_to_excel(existing_file, input_filename)