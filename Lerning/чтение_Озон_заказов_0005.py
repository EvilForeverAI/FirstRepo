import openpyxl

input_file = r"F:\Озон TEMP\Наши товары\Таблица всех артикулов 465.xlsx"

book = openpyxl.open(input_file)

# sheet = book.active # для работы с первым листом, лучше второй вариант
sheet = book.worksheets[0]

# print(sheet.max_row)

# print(sheet[1][0].value)

for row in range(1,sheet.max_row+1):
#for row in sheet.iter_rows(min_row=1, max_row=10, values_only=True):
    print(row)

# if row == 1:
#     v1 = sheet[row][0].value
# else:
#     v1 = int(sheet[row][0].value)
#
# v2 = sheet[row][1].value
# v3 = sheet[row][2].value
# v4 = sheet[row][3].value
#
# print(f"{row}, {v1}, {v2}, {v3}, {v4}")
