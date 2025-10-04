import openpyxl

input_file_1 = r"F:\Озон TEMP\Наши товары\2025.05.30\Аналитика-графики-по всем уровнем категорий — копия2.xlsx"
input_file_2 = r"F:\Озон TEMP\Наши товары\2025.05.30\Аналитика-графики-по всем уровнем категорий — копия.xlsx"

def find_last_column(sheet):
    last_column = 1  # Минимальный столбец (A)
    for row in sheet.iter_rows():
        for cell in reversed(row):
            if cell.value is not None:
                if cell.column > last_column:  # Нашли столбец правее
                    last_column = cell.column
                break  # Переходим к следующей строке
    return last_column

wb1 = openpyxl.load_workbook(input_file_1, read_only=True)  # Основная таблица
wb2 = openpyxl.load_workbook(input_file_2, read_only=True)

sheet1 = wb1.worksheets[0]
sheet2 = wb2.worksheets[0]

print(sheet1[1][-1].coordinate)
print(sheet1.max_column)
print(sheet1[1][0].value)
print(find_last_column(sheet1))

for row in sheet1.iter_rows(max_row=1, values_only=True):
    print(row)

print(sheet2[1][-1].coordinate)
print(sheet2.max_column)
print(sheet2[1][0].value)
print(find_last_column(sheet2))

for row in sheet2.iter_rows(max_row=1, values_only=True):
    print(row)

# for i in range(1,10):
#     for j in range(1,10):
#         print(sheet1.cell(row=i,column=j).value,end=" ")
#     print()

# print(sheet1.cell(row=1,column=3).value)

# print(sheet1.cell(row=1,column=2).value)
# for row_index,row in enumerate(sheet1.iter_rows(max_row=10),start=1):
#     for cell_index, cell in enumerate(row,start=1):
#         print(sheet1.cell(row=row_index,column=cell_index).value,end=" ")
#     print()
# print(find_last_column(sheet1))
# print(sheet1.cell(9,13).value)
